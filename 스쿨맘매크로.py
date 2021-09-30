#스쿨맘 매크로
import openpyxl
wb = openpyxl.load_workbook('schmomtext.xlsx')
sheet1 = wb.active
classnum=input("(1, 2와 같이 숫자만 입력할 것 )/// 학년을 입력해주세요: ")
classnum2=input("(1, 2와 같이 숫자만 입력할 것 )/// 반을 입력해주세요: ")
classroom=input("(정보, 과학B)/// 수업을 입력하세요 : ")
classtime=input("(1, 2와 같이 숫자만 입력할 것 )/// 수업 교시를 입력하세요 : ")
temid=0
stuid=list()
#초기화
for i in range(2,30):
    for j in range(1,8):
        sheet1.cell(row=i, column=j).value=None
    
    
print(("(1, 2와 같이 숫자만 입력할 것 ) 다 입력시에는 t입력 후 엔터"))

while(temid!='t'):
    temid=input("학생 번호를 입력해주세요 : ")
    stuid.append(temid)

for i in range(len(stuid)-1):
    sheet1.cell(row=i+2, column=1).value=classnum
    sheet1.cell(row=i+2, column=2).value=classnum2
    sheet1.cell(row=i+2, column=3).value=stuid[i]
    sheet1.cell(row=i+2, column=4).value=classroom
    sheet1.cell(row=i+2, column=5).value='담당교사입니다. '
    sheet1.cell(row=i+2, column=6).value=classtime+'교시'
    sheet1.cell(row=i+2, column=7).value='수업이 시작되었으니 즉시 수강 바랍니다.'

wb.save('schmomtext.xlsx')
